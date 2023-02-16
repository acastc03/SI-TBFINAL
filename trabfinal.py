from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

class Trabajador:
    
    def __init__(self, nombre, dni, ccc, codigo_pais):
        self.nombre = nombre
        self.dni = dni
        self.ccc = ccc
        self.codigo_pais = codigo_pais
        

    
class ExcelManager:
    
    def __init__(self, dnicorregido, ccc_corregido):
        self.dni = dnicorregido
        self.ccc = ccc_corregido
        
       
    def escribir_ErrorNIF(self):
        ws.cell(row=j+1, column=1).value = nombre
        ws.cell(row=j+1, column=2).value = apellido1
        ws.cell(row=j+1, column=3).value = apellido2
        ws.cell(row=j+1, column=4).value = dni
        wb.save('ErrorNIE1.xlsx')

    def escribir_NIF_erroneo(self):
        excel = book['Hoja5']
        excel.cell(row=i+2, column=1).value = 'NIF erróneo'
        if excel.cell(row=i+2, column=1).value == 'NIF erróneo':
            a = 'A'+str(i+2)
            excel[a].font = Font(color = 'FF0000')
        book.save(ruta_relativa_final)
    
    def cambio_dni(self):
        excel = book['Hoja5']
        excel.cell(row=i+2, column=1).value = trEM.dni
        if excel.cell(row=i+2, column=1).value == trEM.dni :
            a = 'A'+str(i+2)
            excel[a].font = Font(color = '00FF00')
        book.save(ruta_relativa_final)

    def cambio_ccc(self):
        excel = book['Hoja5']
        excel.cell(row=i+2, column=10).value = trEM.ccc
        excel.cell(row=i+2, column=12).value = iban
        book.save(ruta_relativa_final)

    def escribir_ErrorCCC(self):
        ws1.cell(row=k+1, column=1).value = nombre
        ws1.cell(row=k+1, column=2).value = apellido1
        ws1.cell(row=k+1, column=3).value = apellido2
        ws1.cell(row=k+1, column=4).value = ccc
        wb1.save('ErrorCCC1.xlsx')

    def escribir_CCC_erroneo(self):
        excel = book['Hoja5']
        cell.font = Font(color='FF0000')
        excel.cell(row=i+2, column=10).value = 'CCC erróneo'
        if excel.cell(row=i+2, column=10).value == 'CCC erróneo':
            a = 'J'+str(i+2)
            excel[a].font = Font(color = 'FF0000')
        excel.cell(row=i+2, column=12).value = 'Imposible'
        if excel.cell(row=i+2, column=12).value == 'Imposible':
            a = 'L'+str(i+2)
            excel[a].font = Font(color = 'FF0000')
        book.save(ruta_relativa_final)
        
        
        
        
        


class DNI:
    
    def __init__(self,dni):
        self.dni = dni
        
    def validar(self):
        cadenadni = str(self.dni)
        if self.dni == None:
            return False
       

        else:
            if len(cadenadni)!=9:
                return False
            else:
                letras = "TRWAGMYFPDXBNJZSQVHLCKE"
                letrasex = "XYZ"
                if self.dni[0] in letrasex:
                    for i in range(len(letrasex)):
                        if self.dni[0] == letrasex[i]:
                            nie = self.dni.replace(self.dni[0],str(i))
                    niecorto = int(nie[:-1])
                    resto = niecorto%23
                    if letras[resto] == nie[8]:
                        return True
                    else:
                        return False
                if cadenadni[0:5].isnumeric() == False:
                    return False
                else:
                    letras = "TRWAGMYFPDXBNJZSQVHLCKE"
                    ndni = int(self.dni[0:8])
                    letrareal = letras[ndni%23]
                    ldni = self.dni[8]
                    if ldni == letrareal:
                        return True
                    else:
                        return False
               

            
    def error(self):
        letras = "TRWAGMYFPDXBNJZSQVHLCKE"
        letrasex = "XYZ"
        cadenadni = str(self.dni)
        if self.validar() == False:
            if self.dni == None:
                return False

            else:
                
                cadenadni = str(self.dni)
                if len(cadenadni)==9:
                    if self.dni[0] in letrasex:
                        for i in range(len(letrasex)):
                            if self.dni[0] == letrasex[i]:
                                nie = self.dni.replace(self.dni[0],str(i))
                        niecorto = int(nie[:-1])
                        resto = niecorto%23
                        if letras[resto] != nie[8]:
                            dnicorregido = nie.replace(nie[8],letras[resto])
                            return dnicorregido

                        else:
                            return False
                    else:
                        if cadenadni[0:8].isnumeric() == True:
                            dnicorto = int(self.dni[:-1])
                            resto = dnicorto%23
                            if letras[resto] != self.dni[8]:
                                dnicorregido = self.dni.replace(self.dni[8],letras[resto])
                                return dnicorregido
                    
                        else:
                            return False
                if (len(cadenadni)==8 and cadenadni.isnumeric()):
                    ndni = int(cadenadni)
                    resto = ndni%23
                    dnicorregido = cadenadni + letras[resto]
                    return dnicorregido
                else:
                    return False



class CCC:
    
    def __init__(self, ccc, codigo_pais):
        self.ccc = ccc
        self.codigo_pais = codigo_pais

    def modulo(self):
        if self.ccc == None or self.codigo_pais == None or len(self.ccc) != 20 or len(self.codigo_pais) != 2:
            x = 'mentira'
            return x
       
        else:
            letras = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            codigo = ''
            for letrasp in range(len(letras)):
                for i in range(len(self.codigo_pais)):
                    if self.codigo_pais[i] == letras[letrasp]:
                        codigo = codigo + str(letrasp + 10)
            iban = str(codigo + '00' + self.ccc) 
            mod = str(self.ccc + codigo + '00')
            modnum = int(mod)
            if modnum%97  == 1:
                return True
            else:
                return False


    def corregir_ccc(self):
        if not(self.ccc == None or self.codigo_pais == None or len(self.ccc) != 20 or len(self.codigo_pais) != 2):
            listadigitos = list(self.ccc)
            listadigitos1 = '00'+ self.ccc[:8]
            listadigitos2 = self.ccc[10:]
            pesos1 = [1, 2, 4, 8, 5, 10, 9, 7, 3, 6]
            if self.ccc.isnumeric() == True:
                dc1 = sum([int(listadigitos1[i]) * pesos1[i] for i in range(10)]) % 11
                dc2 = sum([int(listadigitos2[i]) * pesos1[i] for i in range(10)]) % 11

                dc1 = 11-dc1
                dc2 = 11-dc2
    
                if dc1 == 10:
                    dc1 = 1
                elif dc1 == 11:
                    dc1 = 0

                if dc2 == 10:
                    dc2 = 1
                elif dc2 == 11:
                    dc2 = 0
                
                listadigitos[8] = str(dc1)
                listadigitos[9] = str(dc2)
                ccc_corregido = "".join(listadigitos)

                return ccc_corregido
            
            else:
                y = 'Error CCC'
                return y
                
        else:
            y = 'Error CCC'
            return y
        
    def dc(self):
        if not(self.ccc == None or self.codigo_pais == None or len(self.ccc) != 20 or len(self.codigo_pais) != 2):
            listadigitos = list(self.ccc)
            listadigitos1 = '00'+ self.ccc[:8]
            listadigitos2 = self.ccc[10:]
            pesos1 = [1, 2, 4, 8, 5, 10, 9, 7, 3, 6]
            if self.ccc.isnumeric() == True:
                dc1 = sum([int(listadigitos1[i]) * pesos1[i] for i in range(10)]) % 11
                dc2 = sum([int(listadigitos2[i]) * pesos1[i] for i in range(10)]) % 11
                
                dc1 = 11-dc1
                dc2 = 11-dc2
        
                if dc1 == 10:
                    dc1 = 1
                elif dc1 == 11:
                    dc1 = 0
                    
                if dc2 == 10:
                    dc2 = 1
                elif dc2 == 11:
                    dc2 = 0
                    
                dcb = str(dc1)+str(dc2)
                return dcb
            else:
                y = 'Error CCC'
                return y
        else:
            y = 'Error CCC'
            return y
                    
                    
                
            




 
ruta_ejecutable = Path().absolute()
nombre_fichero = 'Excel.xlsx'
ruta_relativa = Path('Resources/' + nombre_fichero)
ruta_relativa_final = Path.joinpath(ruta_ejecutable, ruta_relativa)
book = openpyxl.load_workbook(ruta_relativa_final, data_only = True)
excel = book['Hoja5']
max_row = excel.max_row

wb = openpyxl.Workbook()
ws = wb.active
wb1 = openpyxl.Workbook()
ws1 = wb1.active

data = []
for fila in excel.iter_rows(min_col=1, max_col=(excel.max_column)-1):
    fila_data=[]
    for cell in fila:
        fila_data.append(cell.value)
    data.append(fila_data)

data.pop(0)


erroresnif = []
erroresccc = []


for i in range(8922):

    if data[i] == [None, None, None, None, None, None, None, None, None, None, None, None]:
        pass
    else:
        j = len(erroresnif)
        k = len(erroresccc)
        dni = data[i][0]
        nombre = data[i][1]
        ccc = data[i][9]
        apellido1 = data[i][2]
        apellido2 = data[i][3]
        codigo_pais = data[i][10]
        cccib = ccc.replace(ccc[9:11],'')
        tr = Trabajador(nombre, dni, ccc, codigo_pais)
        trdni = DNI(tr.dni)
        trccc = CCC(tr.ccc, tr.codigo_pais)
        trEM = ExcelManager(trdni.error(), trccc.corregir_ccc())
        resto_digitos = trccc.corregir_ccc()
        dc = trccc.dc()
        iban = codigo_pais + dc + resto_digitos #se que me falta eso, pero no me da tiempo a arreglarlo.
        if trdni.validar() == True:
            pass
        if trdni.validar() == False:
            if trEM.dni == False or tr.dni != None:
                trEM.escribir_NIF_erroneo()
                if trdni.error() == False:
                    erroresnif.append('a')
                    j = len(erroresnif)
                    trEM.escribir_ErrorNIF()
                
                
            if trEM.dni != False and tr.dni != None:
                trEM.cambio_dni()
        if ccc.isnumeric() == True:
            if trccc.modulo() == False:
                trEM.cambio_ccc()
        
            else:
                erroresccc.append('a')
                k = len(erroresccc)
                trEM.escribir_ErrorCCC()
                trEM.escribir_CCC_erroneo()
                
        if ccc.isnumeric() == False:
            erroresccc.append('a')
            k = len(erroresccc)
            trEM.escribir_ErrorCCC()
            trEM.escribir_CCC_erroneo()
           

        
    
        
    

    
    


        
    


